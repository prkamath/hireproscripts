from dbconstants import *
import MySQLdb
from openpyxl import load_workbook
from datetime import datetime,timedelta
import uuid
import json
import requests
import random
import copy
import sys,os
import traceback

def checkForCandidates(candidatelist):
    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()
    totalexistingcandidates=0
    print "\n===========================================\n"
    print "The emails we will test for are",candidatelist
    print "\n===========================================\n"
    for emailid in candidatelist:
        fetched=0
        tempString="select id,email1 from candidates where email1='%s'"%emailid
        cur.execute(tempString)
        for row in cur.fetchall():
            print row
            fetched=1
        if (fetched==1):
            totalexistingcandidates=totalexistingcandidates+1
        else:
            print "Unable to fetch record for " + emailid
    print "Total fetched candidates are " + str(totalexistingcandidates) 
    db.close()

def parseCallStatusDetails(stringVal,callStatusDetails):
    allDetails=stringVal.splitlines()
    for detail in allDetails:
        detail=detail.encode("utf-8")
        detail=detail.replace('\xe2\x80\x93','-')
        detail=detail.replace(':','-')
        commentDate=detail.split("-")[0]
        print commentDate
        newcommentDate=makeSenseOfDate(commentDate)
        print newcommentDate
        newcommentDate=newcommentDate.replace("th","")
        try:
            finalDate=datetime.strptime(newcommentDate,"%d %b %Y")
        except:
            continue

        singleConvrsn={}
        singleConvrsn['date']=finalDate
        singleConvrsn['comment']=detail.split("-")[1]
        callStatusDetails.append(singleConvrsn)

def checkForUsers(userlist):
    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()
    totalexistingusers=0
    print "\n===========================================\n"
    print "The emails we will test for are",userlist
    print "\n===========================================\n"
    for emailid in userlist:
        fetched=0
        tempString="select id,email1 from users where email1='%s'"%emailid
        cur.execute(tempString)
        for row in cur.fetchall():
            print row
            fetched=1
        if (fetched==1):
            totalexistingusers=totalexistingusers+1
        else:
            print "Unable to fetch record for " + emailid
    print "Total fetched users are " + str(totalexistingusers) 
    db.close()

def makeSenseOfDate(inDate):
    if (inDate.lower().find("jan")>-1 or inDate.lower().find("feb")>-1):
        outDate=inDate + " 2016"
        return outDate
    outDate=inDate+" 2015"
    return outDate

def morphRows(allRows):
    for singleRow in allRows:
        singleRow['EmailId']="%s%d"%(singleRow['EmailId'],random.randint(1,20000))
        singleRow['CandidateId']=singleRow['CandidateId']+random.randint(1,10000)

def parseRowIntoDict(row,singleRow):
    if (row[0].value==None):
        return -1
    singleRow['CandidateId']=row[0].value#Candidate Id
    singleRow['Name']=row[1].value #Candidate Full Name
    singleRow['Mobile1']=row[2].value #Candidate Mobile Phone
    singleRow['EmailId']=row[3].value
    singleRow['Level']=row[4].value # Level
    singleRow['BU']=row[5].value # BU
    singleRow['OwningDepartment']=row[6].value # Owning Department
    singleRow['ActualWorkLocation']=row[7].value # Actual Work Location (City) 
    singleRow['PrimaryRecruiterName']=row[8].value #Primary Recruiter Name
    singleRow['PrimarySkills']=row[9].value #Primary Skills
    singleRow['TentativeDOJ']=row[11].value #C Hire Date
    singleRow['ExpectedDOJ']=row[13].value
    singleRow['LastDateCalled']=row[15].value
    singleRow['Action']=row[17].value
    singleRow['Inconsistencies']=[]
    singleRow['DeclinedReason']=row[28].value# Declined Reason
    singleRow['LikeToNegotiate']=row[29].value# Like to Negotiate (Declined cases).. id = 142604 (NAVEENA)
    singleRow['spocname']=row[30].value# POFU SPOC
    tempQueryDetails={}
    singleRow['QueryDetails']=tempQueryDetails
    try:
        q_type = row[19].value
        if q_type != 'NA' and q_type != '':
            tempQueryDetails['QueryLevelRaised']=row[18].value
            tempQueryDetails['QueryType'] = q_type
            tempQueryDetails['QueryRaisedDate']=row[20].value
            tempQueryDetails['QueryResolvedDate']=row[21].value
            tempQueryRemarks=row[24].value#Query Remarks
            tempQueryDetails['QueryRemarks']=tempQueryRemarks.encode('latin-1', 'ignore')
            #gap=tempQueryDetails['QueryResolvedDate']-tempQueryDetails['QueryRaisedDate']
            #tempQueryDetails['NoOfDaysForQueryResolution']=gap.days
    except:
        singleRow['Inconsistencies'].append("DateTimeError")

    singleRow['Stage']=row[25].value#Stages
    singleRow['JoiningStatus']=row[26].value#Status As Per Last Call

    callStatusDetails=[]
    singleRow['CallStatus']=callStatusDetails
    tempustring=row[27].value#Remarks
    singleRow['AllCallDetails']=tempustring.encode('latin-1', 'ignore')
    #parseCallStatusDetails(row[27].value,callStatusDetails)
    return 0


def getCandidateStaffingProfileId(third_party_id):
    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()

    tempString="select candidatestaffingprofile_id from candidates where third_party_id=%d and tenant_id=%d"%(third_party_id,DB_TENANT_ID)
    cur.execute(tempString)
    row=cur.fetchone()
    if (row == None):
        print "Unable to get CSP_ID!! for %d"%third_party_id
        return -1
    candId=row[0]
    db.close()
    return candId

def getCandidateId(third_party_id):
    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()

    tempString="select id from candidates where third_party_id='%d' and tenant_id=%d"%(third_party_id,DB_TENANT_ID)
    cur.execute(tempString)
    row=cur.fetchone()
    if (row == None):
        print "Unable to get CA_ID!! for %d"%third_party_id
        return -1
    candId=row[0]
    db.close()
    return candId

def entryForCandSpocExists(candidateid,spocName,spocDict):
    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()
    userid=spocDict[spocName]
    tempString="select id from candidate_spocs where candidate_id=%d and user_id=%d"%(candidateid,userid)
    cur.execute(tempString)
    row=cur.fetchone()
    if (row != None):
        print "Entry for candidate exists"
        return 1
    else:
        return 0

def createCandSpocs(candidateid,spocName,spocDict):
    if (entryForCandSpocExists(candidateid,spocName,spocDict)):
        print "Repeated entry"
        return 1

    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()
    guidval=uuid.uuid4()
    tempstring="insert into candidate_spocs (version,candidate_id,user_id,guid,is_deleted,created_on,created_by) values (0,%d,%d,'%s',0,now(),%d)"\
               %(candidateid,spocDict[spocName],guidval,SPOC_CREATED_BY)
    if VERBOSE_DEBUG_SETTING:
        print tempstring
    cur.execute(tempstring)
    db.commit()
    db.close()
    return 1


def updateCandidateStaffingProfile(staffingProfileId,expectedDateOfJoining):
    return # NAVEENA
    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()
    dateOfJ=expectedDateOfJoining.strftime('%Y-%m-%d')
    print dateOfJ
    tempString="update candidate_staffing_profiles set expected_joining_date='%s' where id=%d"%(dateOfJ,staffingProfileId)
    if VERBOSE_DEBUG_SETTING:
        print tempString
    cur.execute(tempString)
    db.commit()
    db.close()

def int_executeQuery(query,queryParamList):
    resultSet = []
    columns = {}
    inserted_id = None
    try:
        db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
        cursor = db.cursor()
        cursor.execute(query,queryParamList)
        inserted_id = cursor.lastrowid
        resultSet = cursor.fetchall()
        desc = cursor.description
        if desc:
            for i in range(len(desc)):
                columns[desc[i][0]] = i
        db.commit()
        cursor.close()
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
	print("Error in creating/executing query against DB [%s] with params %s [%s:%s:%s]" %(query,queryParamList,fname,exc_tb.tb_lineno,str(e)))
    return (columns,resultSet,inserted_id)

def createCallData(ca_id,spoc,created_on,remarks):
    query = """insert into staffing_pofus (version, candidate_id,tenant_id,created_by,created_on, is_deleted,guid)
             values (%s,%s,%s,%s,%s,%s,%s)"""
    qplist = []
    sp_id = None
    qplist.append("1")
    qplist.append(ca_id)
    qplist.append(DB_TENANT_ID)
    qplist.append(spoc)
    qplist.append(str(created_on))
    qplist.append("0")
    qplist.append(str(uuid.uuid4()))
    (colIdx,resultSet,sp_id) = int_executeQuery(query,qplist)
    if VERBOSE_DEBUG_SETTING:
        print ("Inserted SP entry for cand=%s with id =%s"%(ca_id,sp_id))

    query = """insert into staffing_pofu_calls (caller, call_type,current_status_id,created_by,created_on,staffingpofu_id)
             values (%s,%s,%s,%s,%s,%s)"""
    qplist = []
    qplist.append(spoc)
    qplist.append(OUTBOUND_CALL)
    qplist.append(COMPLETED_CALL)
    qplist.append(spoc)
    qplist.append(str(created_on))
    qplist.append(sp_id)
    (colIdx,resultSet,spc_id) = int_executeQuery(query,qplist)
    if VERBOSE_DEBUG_SETTING:
        print ("Inserted SPC entry for cand=%s with id =%s"%(ca_id,spc_id))
    query = """insert into staffing_pofu_call_historys (status_id,comment,created_by,created_on,staffingpofucall_id) 
               values(%s,%s,%s,%s,%s)
            """
    qplist = []
    qplist.append(COMPLETED_CALL)
    qplist.append(remarks)
    qplist.append(spoc)
    qplist.append(str(created_on))
    qplist.append(spc_id)
    (colIdx,resultSet,spch_id) = int_executeQuery(query,qplist)
    if VERBOSE_DEBUG_SETTING:
        print ("Inserted SPCH entry for cand=%s with id =%s"%(ca_id,spch_id))

def createStatusEntries(ca_id,new_status_id,spoc,last_called,ready_to_negotiate,declinereason_id):
    """
        qplist.append(ca_id)
        qplist.append(new_status_id)
        qplist.append(spoc)
        qplist.append(last_called)
    mysql> desc staffing_statuss;
    +--------------------+-------------+------+-----+---------+----------------+
    | Field              | Type        | Null | Key | Default | Extra          |
    +--------------------+-------------+------+-----+---------+----------------+
    | id                 | int(11)     | NO   | PRI | NULL    | auto_increment |
    | version            | int(11)     | NO   |     | NULL    |                |
    | current_status_id  | int(11)     | NO   |     | NULL    |                |
    | responsiveness     | int(11)     | YES  |     | NULL    |                |
    | observations       | int(11)     | YES  |     | NULL    |                |
    | resignation_status | int(11)     | YES  |     | NULL    |                |
    | decline_reason     | int(11)     | YES  |     | NULL    |                |
    | comments           | mediumtext  | YES  |     | NULL    |                |
    | tenant_id          | int(11)     | YES  | MUL | NULL    |                |
    | candidate_id       | int(11)     | YES  |     | NULL    |                |
    | created_by         | int(11)     | NO   |     | NULL    |                |
    | created_on         | datetime    | NO   |     | NULL    |                |
    | modified_by        | int(11)     | YES  |     | NULL    |                |
    | modified_on        | datetime    | YES  |     | NULL    |                |
    | is_deleted         | tinyint(1)  | NO   |     | NULL    |                |
    | guid               | varchar(40) | NO   |     | NULL    |                |
    +--------------------+-------------+------+-----+---------+----------------+
    16 rows in set (0.00 sec)
    mysql> desc staffing_status_historys;
    +--------------------+------------+------+-----+---------+----------------+
    | Field              | Type       | Null | Key | Default | Extra          |
    +--------------------+------------+------+-----+---------+----------------+
    | id                 | int(11)    | NO   | PRI | NULL    | auto_increment |
    | status_id          | int(11)    | YES  |     | NULL    |                |
    | comments           | mediumtext | YES  |     | NULL    |                |
    | responsiveness     | int(11)    | YES  |     | NULL    |                |
    | observations       | int(11)    | YES  |     | NULL    |                |
    | resignation_status | int(11)    | YES  |     | NULL    |                |
    | decline_reason     | int(11)    | YES  |     | NULL    |                |
    | created_by         | int(11)    | NO   |     | NULL    |                |
    | created_on         | datetime   | NO   |     | NULL    |                |
    | modified_by        | int(11)    | YES  |     | NULL    |                |
    | modified_on        | datetime   | YES  |     | NULL    |                |
    | staffingstatus_id  | int(11)    | YES  | MUL | NULL    |                |
    +--------------------+------------+------+-----+---------+----------------+
    12 rows in set (0.00 sec)
    """
    create_ss_entry = True
    create_ssh_entry = True
    # Get the current status and if its different from the current, create an entry in ss_historys and udpates ss
    query = "select id ,version, current_status_id from staffing_statuss where tenant_id = %s and candidate_id = %s"
    qplist = []
    qplist.append(DB_TENANT_ID)
    qplist.append(ca_id)
    ss_id = None
    version = 0
    resultSet = []
    colIdx    = {}
    unused = None
    is_deleted=0
    (colIdx,resultSet,unused) = int_executeQuery(query,qplist)
    if len(resultSet) > 0:
        ss_id = resultSet[0][colIdx["id"]]
        if VERBOSE_DEBUG_SETTING:
            print ("SS entry already exists for cand=%s with id =%s"%(ca_id,ss_id))
        version = resultSet[0][colIdx["version"]]
        create_ss_entry = False
        if resultSet[0][colIdx["current_status_id"]] == new_status_id:
            print ("CurrentStatus for cand=%s with id =%s is the same as the new one"%(ca_id,ss_id))
            create_ssh_entry = False
    if create_ss_entry == True:
        query = """insert into staffing_statuss (version,current_status_id,tenant_id,candidate_id,created_by,created_on,is_deleted,guid,decline_reason,ready_to_negotiate) 
                   values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        qplist = []
        qplist.append("1")
        qplist.append(new_status_id)
        qplist.append(DB_TENANT_ID)
        qplist.append(ca_id)
        qplist.append(spoc)
        qplist.append(last_called)
        qplist.append("0")
        qplist.append(str(uuid.uuid4()))
        qplist.append(declinereason_id)
        qplist.append(ready_to_negotiate)
        (colIdx,resultSet,ss_id) = int_executeQuery(query,qplist)
        if VERBOSE_DEBUG_SETTING:
            print ("Inserted SS entry for cand=%s with id =%s"%(ca_id,ss_id))
    else:
        query = """ update staffing_statuss set version = %s,current_status_id=%s,modified_by=%s,modified_on=%s,decline_reason=%s,ready_to_negotiate=%s where id = %s"""
        version += 1
        qplist = []
        qplist.append(version)
        qplist.append(new_status_id)
        qplist.append(spoc)
        qplist.append(str(last_called))
        qplist.append(declinereason_id)
        qplist.append(ready_to_negotiate)
        qplist.append(ss_id)
        int_executeQuery(query,qplist)
        if VERBOSE_DEBUG_SETTING:
            print ("Updated SS entry for id =%s"%(ss_id))

    if create_ssh_entry == True:
        #First update the older entry
        olderDate=last_called - timedelta(2)#2 days in the past
        olderDate=str(olderDate)
        query = """update staffing_status_historys set created_on = %s where staffingstatus_id=%s"""
        qpList = []
        qpList.append(olderDate)
        qpList.append(ss_id)

        int_executeQuery(query,qpList)
        if VERBOSE_DEBUG_SETTING:
            print ("Updated history entry for id =%s"%(ss_id))


        query = """insert into staffing_status_historys(status_id,
                   created_by,created_on,staffingstatus_id,decline_reason)
                   values(%s,%s,%s,%s,%s)"""
        qplist = []
        qplist.append(new_status_id)
        qplist.append(spoc)
        qplist.append(str(last_called))
        qplist.append(ss_id)
        qplist.append(declinereason_id)

        sshid = None
        (colIdx,resultSet,sshid) = int_executeQuery(query,qplist)
        if VERBOSE_DEBUG_SETTING:
            print ("Inserted SSH entry with id =%s"%(sshid))

def insertCandStaffingQueries(csp_id,q_cat,q_criticality,spoc,created_on,resolved_on,q_details,is_pending):
    """
    +------------------------+----------+------+-----+---------+----------------+
    | Field                  | Type     | Null | Key | Default | Extra          |
    +------------------------+----------+------+-----+---------+----------------+
    | id                     | int(11)  | NO   | PRI | NULL    | auto_increment |
    | tenant_id              | int(11)  | YES  | MUL | NULL    |                |
    | created_by             | int(11)  | NO   |     | NULL    |                |
    | created_on             | datetime | NO   |     | NULL    |                |
    | modified_by            | int(11)  | YES  |     | NULL    |                |
    | modified_on            | datetime | YES  |     | NULL    |                |
    | is_pending             | int(11)  | YES  |     | NULL    |                |
    | candstaffingprofile_id | int(11)  | YES  | MUL | NULL    |                |
    | querycategory_id       | int(11)  | YES  |     | NULL    |                |
    | querycriticality_id    | int(11)  | NO   |     | NULL    |                |
    | query_details          | text     | YES  |     | NULL    |                |
    +------------------------+----------+------+-----+---------+----------------+
    """
    db = MySQLdb.connect(host=DB_IP,    # your host, usually localhost
            user=DB_USER,         # your username
            passwd=DB_PASSWORD,  # your password
            db=DB_DBNAME)        # name of the data base
    cur = db.cursor()
    query = """ insert into cand_staffing_querys (tenant_id,created_by,created_on,modified_by,modified_on,
                is_pending,candstaffingprofile_id, 
                querycategory_id, querycriticality_id,query_details) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
    q_paramlist = []
    q_paramlist.append(DB_TENANT_ID)
    q_paramlist.append(spoc)
    q_paramlist.append(created_on)
    q_paramlist.append(spoc)
    q_paramlist.append(resolved_on)
    q_paramlist.append(is_pending)
    q_paramlist.append(csp_id)
    q_paramlist.append(q_cat)
    q_paramlist.append(q_criticality)
    q_paramlist.append(q_details)
    if VERBOSE_DEBUG_SETTING:
        print query
        print q_paramlist
    cur.execute(query,q_paramlist)
    db.commit()
    db.close()

def createCandidates(allRows):
    allDicts=[]
    dict1=json.loads(template)
    candidateList=dict1["candidateList"]

    for row in allRows:
        tempDict=copy.deepcopy(candidateList[0])
        tempDict["OriginalSourceID"]=row['CandidateId']
        tempDict["Email1"]=row['EmailId']
        tempDict["Name"]=row['Name']
        tempDict["Mobile1"]=str(random.randint(1,99999999999))
        allDicts.append(tempDict)

    dict1["candidateList"]=allDicts
    jsonStr=json.dumps(dict1)
    print jsonStr
    serviceUrl=SERVICE_URL%(SERVICE_IP)
    headers={}
    headers['Content-Type']="""application/json; charset=utf-8"""
    headers['X-AUTH-TOKEN']=AUTH_TOKEN
    ret=requests.post(url=serviceUrl,data=jsonStr,headers=headers)
    print ret.text

def checkSanityOfDicts(singleRow,spocdict,query_cat_dict,query_criticality_dict,status_dict,rowid):
    spocname=singleRow["spocname"]
    if (spocname  not in spocdict):
        print "####Missing spocname  " + spocname + str(rowid)
        raise ValueError('A very specific bad thing happened')

    if 'QueryType' in singleRow['QueryDetails']:
        querytype=singleRow['QueryDetails']['QueryType']
        if (querytype not in query_cat_dict):
            print "#####Missing query category#####   "  + str(rowid)
            raise ValueError('A very specific bad thing happened')

    stage = singleRow["Stage"]
    if (stage.lower() not in status_dict):
        print "#####Missing Stage!!#####  " + str(rowid)
        raise ValueError('A very specific bad thing happened')
    else:
        valid_status_dict = status_dict.get(stage.lower(),None)
    joiningstatus=singleRow["JoiningStatus"]
    if (joiningstatus.lower() not in valid_status_dict):
        print "#####Missing Status in the given status!! #####  " + str(rowid) + " " + stage + " " + joiningstatus
        print  valid_status_dict
        print "\n\n\n"
        print singleRow
        raise ValueError('A very specific bad thing happened')

def createAllCandidates(allRows):
    #MorphRows changes all the rows' emailId and candidateId
    morphRows(allRows)
    tempRows=[]
    #Create the candidates also
    for row in allRows:
        tempRows.append(row)
        if (len(tempRows)==MAX_ROWS_TO_CREATE):
            createCandidates(tempRows)
            tempRows=[]

    createCandidates(tempRows)

if __name__=="__main__":
    #candId=getCandidateStaffingProfileId("prkamath@gmail.com")
    #updateCandidateStaffingProfile(candId,datetime(2016, 1, 28, 0, 0))
    #print candId
    wb2=load_workbook(XCEL_SHEET_NAME)
    sheetnames=wb2.get_sheet_names()
    ws=wb2.active
    count=0
    allRows=[]
    spocdict={}
    query_cat_dict={}
    query_criticality_dict={}
    status_dict={}
    populateMetaData(spocdict,query_cat_dict,query_criticality_dict,status_dict)

    count=0

    print "\n\n\n\n"
    for row in ws.iter_rows(row_offset=1):
        singleRow={}
        ret=parseRowIntoDict(row,singleRow)
        if (0==ret) and (count in range(START_ROW_TO_PARSE,START_ROW_TO_PARSE+MAX_ROWS_TO_PARSE)) and ('EmailId' in singleRow):
            allRows.append(singleRow)
        if (0==ret):
            checkSanityOfDicts(singleRow,spocdict,query_cat_dict,query_criticality_dict,status_dict,count)
        count = count+1


    #Here we ensure uniqueness of the createdCandidates. This and the next line is purely for debuggin
    #When actually using we know that the candidates would have already been created
    #createAllCandidates(allRows)

    #Load all remaining pieces
    for singleRow in allRows:
        #print singleRow
        if ('EmailId' in singleRow):
            singleRow['candidatestaffingprofileid']= getCandidateStaffingProfileId(singleRow['CandidateId'])
            singleRow['CandidateIdPrimaryKey']=getCandidateId(singleRow['CandidateId'])

    for singleRow in allRows:
        if (singleRow['CandidateIdPrimaryKey'] == -1):
            continue
        #print singleRow
        try:
            #Now update the DOJ for all entries
            updateCandidateStaffingProfile(singleRow['candidatestaffingprofileid'],singleRow['ExpectedDOJ'])

            #If we cannot create a spocs then we will not proceed
            if (not createCandSpocs(singleRow['CandidateIdPrimaryKey'],singleRow['spocname'],spocdict)):
                continue
            is_pending = 0
            queryDetails=singleRow['QueryDetails']

            if 'QueryType' in queryDetails:
                query_resolved_on = queryDetails['QueryResolvedDate']
                invalidDate=datetime(1899, 12, 30, 0, 0)
                if queryDetails['QueryResolvedDate'] == invalidDate :
                    is_pending = 1
                    query_resolved_on = None
            
                #insertCandStaffingQueries(csp_id,q_cat,q_criticality,spoc,created_on,resolved_on,q_details,is_pending):
                insertCandStaffingQueries(singleRow['candidatestaffingprofileid'],
                            query_cat_dict[queryDetails['QueryType']],
                            query_criticality_dict[queryDetails['QueryLevelRaised']],
                            spocdict[singleRow['spocname']],
                            queryDetails['QueryRaisedDate'],
                            query_resolved_on,
                            queryDetails['QueryRemarks'],
                            is_pending
                            )
            tmpNegotiate = singleRow['LikeToNegotiate']
            ready_to_negotiate = None
            if tmpNegotiate:
                if tmpNegotiate.lower() == 'yes':
                    ready_to_negotiate = "1"
                elif tmpNegotiate.lower() == 'no':
                    ready_to_negotiate = "0"
            declinereason_id = None
            relevant_status_dict = status_dict.get(singleRow["Stage"].lower(),None)
            if relevant_status_dict:
                stat_n_reasons = relevant_status_dict.get(singleRow["JoiningStatus"].lower(),None)
                new_status_id = stat_n_reasons["id"]
                reason = singleRow['DeclinedReason']
                if reason:
                    if reason.lower() != 'na' and reason.lower() != '' and reason.lower() != None:
                        reasons_dict = stat_n_reasons["reasons"]
                        declinereason_id = reasons_dict.get(reason.lower(),None) 
                        if declinereason_id == None:
                            print "Declined reason %s not present in the dict!!"%(reason)
                            print stat_n_reasons
                            print singleRow
            createStatusEntries(
                                singleRow['CandidateIdPrimaryKey'],
                                new_status_id,
                                spocdict[singleRow['spocname']],
                                singleRow['LastDateCalled'],
                                ready_to_negotiate,
                                declinereason_id
                            )
            createCallData(
                        singleRow['CandidateIdPrimaryKey'],
                        spocdict[singleRow['spocname']],
                        singleRow['LastDateCalled'],
                        singleRow['AllCallDetails'])

        except:
            tb = traceback.format_exc()
            print tb
    print "Total rows=" + str(len(allRows))

'''
    checkForUsers(["prkamath@gmail.com","sheela@we.com"])
    checkForCandidates(["prkamath@gmail.com","sheela@we.com"])
'''
